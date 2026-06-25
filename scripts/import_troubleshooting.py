#!/usr/bin/env python3
"""
Import the troubleshooting master spreadsheet into the site's data file.

  Reads : data/troubleshooting-master.xlsx  (the master you edit)
  Writes: js/troubleshooting-data.js         (what the website loads)

The "Internal note" column is dropped here — it never reaches the website.
Run after editing the spreadsheet:

  .venv/bin/python scripts/import_troubleshooting.py
"""
import json
import sys
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "troubleshooting-master.xlsx"
OUT = ROOT / "js" / "troubleshooting-data.js"

# Customer-facing "Then…" value  ->  flow code used by the page
OUTCOMES = {
    "solved": "solved",
    "next step": "next",
    "contact support": "support",
}

# Column positions in the master sheet (1-based)
(C_ID, C_CAT, C_ISSUE, C_STEP, C_CHECK, C_RESULT, C_SHOW,
 C_IMAGE, C_THEN, C_EVIDENCE, C_INTERNAL) = range(1, 12)
NCOLS = 11

# Manual illustration: base asset name -> {dark, light} sources + alt text source.
IMAGE_DIR = "assets/manual"


def clean(v):
    return "" if v is None else str(v).strip()


def main():
    if not SRC.exists():
        sys.exit(f"Master spreadsheet not found: {SRC}")

    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Troubleshooting"] if "Troubleshooting" in wb.sheetnames else wb.active

    issues = {}          # id -> issue dict (preserves insertion order)
    categories = []      # first-appearance order
    errors = []

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, NCOLS + 1)]
        if all(v in (None, "") for v in row):
            continue

        idv = clean(row[C_ID - 1])
        category = clean(row[C_CAT - 1])
        issue = clean(row[C_ISSUE - 1])
        check = clean(row[C_CHECK - 1])
        result = clean(row[C_RESULT - 1])
        show = clean(row[C_SHOW - 1])
        image = clean(row[C_IMAGE - 1])
        then_raw = clean(row[C_THEN - 1])
        evidence = clean(row[C_EVIDENCE - 1])

        if not idv:
            errors.append(f"Row {r}: missing ID")
            continue
        outcome = OUTCOMES.get(then_raw.lower())
        if outcome is None:
            errors.append(f"Row {r} ({idv}): invalid 'Then…' value {then_raw!r} "
                          f"(expected: Solved / Next step / Contact support)")
            continue
        if outcome == "next" and not show:
            errors.append(f"Row {r} ({idv}): 'Next step' needs a fix in 'Show the customer' "
                          f"(it becomes a 'try this' step shown to the customer).")
            continue

        if idv not in issues:
            issues[idv] = {"id": idv, "category": category, "issue": issue, "checks": []}
            if category and category not in categories:
                categories.append(category)

        issue_obj = issues[idv]
        # Group consecutive rows that share the same check text into one question
        check_obj = None
        for c in issue_obj["checks"]:
            if c["question"] == check:
                check_obj = c
                break
        if check_obj is None:
            check_obj = {"question": check, "options": []}
            issue_obj["checks"].append(check_obj)

        option = {
            "label": result,
            "guidance": show,
            "outcome": outcome,
            "evidence": evidence,
        }
        if image:
            # Comma-separated asset base names; optional "@NN" sets a size scale (percent)
            imgs = []
            for raw in image.split(","):
                nm = raw.strip()
                if not nm:
                    continue
                scale = 1.0
                if "@" in nm:
                    nm, pct = nm.split("@", 1)
                    nm = nm.strip()
                    try:
                        scale = float(pct) / 100.0
                    except ValueError:
                        scale = 1.0
                fig = {"dark": f"{IMAGE_DIR}/{nm}.png", "light": f"{IMAGE_DIR}/{nm}-ink.png"}
                if scale != 1.0:
                    fig["scale"] = round(scale, 3)
                imgs.append(fig)
            if outcome == "next":
                # Picture(s) belong to the "try this" fix shown after this option
                option["images"] = imgs
            else:
                # Picture(s) belong to the question itself (helps the customer answer)
                check_obj["images"] = imgs
        check_obj["options"].append(option)

    if errors:
        sys.exit("Import aborted — fix these and re-run:\n  " + "\n  ".join(errors))

    data = {
        "generatedAt": date.today().isoformat(),
        "categories": categories,
        "issues": list(issues.values()),
    }

    payload = json.dumps(data, ensure_ascii=False, indent=2)
    banner = ("/* GENERATED FILE — do not edit by hand.\n"
              "   Source: data/troubleshooting-master.xlsx\n"
              "   Rebuild: .venv/bin/python scripts/import_troubleshooting.py */\n")
    OUT.write_text(f"{banner}window.ENSO_TROUBLESHOOTING = {payload};\n", encoding="utf-8")

    print(f"Wrote {OUT.relative_to(ROOT)}")
    print(f"  issues: {len(data['issues'])}  | categories: {len(categories)}")
    counts = {}
    for iss in data["issues"]:
        for ch in iss["checks"]:
            for op in ch["options"]:
                counts[op["outcome"]] = counts.get(op["outcome"], 0) + 1
    print(f"  outcomes: {counts}")


if __name__ == "__main__":
    main()
